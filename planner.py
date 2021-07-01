import read
import math
import people
import os
from openpyxl import load_workbook, Workbook, workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import binpacking


class Planner():
    """ Main class that does all the actions - search & retrieve from all_data, split into grps, then output into a session obj """

    def __init__(self, elderly_data, volunteer_data):
        self.all_elderly = elderly_data
        self.all_volunteers = volunteer_data

    # Main methods
    def write_elderly_grouping(self, n, filename):
        wb = Workbook()
        ws = wb.active
        row = 1
        ws['A1'] = 'Hosp'
        ws['C1'] = 'Blk'
        ws['D1'] = 'Floor'
        ws.column_dimensions['B'].width = 28
        grouping = self.create_elderly_grouping(n)
        grp = 1
        for group in grouping:
            _ = ws.cell(row=row, column=2, value='Group ' +
                        str(grp) + ':')
            row += 1
            for e in group:
                for col in range(1, 5):
                    if col == 1:
                        _ = ws.cell(row=row, column=col, value=e.hospitability)
                    elif col == 2:
                        name = e.get_name()
                        if e.nickname:
                            name = name + ' (' + e.nickname + ')'
                        _ = ws.cell(row=row, column=col, value=name)
                    elif col == 3:
                        _ = ws.cell(row=row, column=col, value=e.blk)
                    elif col == 4:
                        _ = ws.cell(row=row, column=col, value=e.floor)
                row += 1
            row += 1
            grp += 1
        wb.save(filename)

    def create_session(self, date, grouping_filename):
        return Session(date, self.load_elderly_grouping(grouping_filename), self.get_available(date))

    def suspend(self, elderly):
        """ elderly = str """
        self.all_elderly[self.all_elderly.index(
            elderly)].suspend()

    def unsuspend(self, elderly):
        self.all_elderly[self.all_elderly.index(
            elderly)].unsuspend()

    def remove_volunteer(self, volunteer):
        if type(volunteer) == str:
            volunteer = self.get_person(volunteer)
        self.output_volunteer_detail(
            volunteer, f'Archive/Previous volunteers/{volunteer.get_name()} profile.xlsx')
        self.all_volunteers.remove(volunteer.get_name())
        for v in self.all_volunteers:
            if volunteer.get_name() in v.friend:
                del v.friend[volunteer.get_name()]

    def remove_elderly(self, elderly):
        if type(elderly) == str:
            elderly = self.get_person(elderly)
        self.all_elderly.remove(elderly.get_name())
        for v in self.all_volunteers:
            if elderly.get_name() in v.visited:
                del v.visited[elderly.get_name()]

    def output_volunteer_detail(self, volunteer, filename, all=False):
        """ allow volunteer to input as list for outputting multiple to .xlsx """
        if type(volunteer) == str:
            volunteer = self.get_person(volunteer)
        wb = Workbook()
        ws = wb.active
        ws['A1'] = volunteer.get_name()
        ws['A2'] = 'Total hours ='
        ws['B2'] = len(volunteer.attendance)*2.0
        ws['A3'] = 'Attended'
        for i in range(len(volunteer.attendance)):
            _ = ws.cell(row=3, column=i+2,
                        value=volunteer.attendance[i].date().strftime('%d/%m/%y'))
        if all:
            visited = list(sorted(volunteer.visited.items(),
                                  key=lambda x: x[1], reverse=True))
            friends = list(sorted(volunteer.friend.items(),
                                  key=lambda x: x[1], reverse=True))
        else:
            visited = list(sorted(
                volunteer.visited.items(), key=lambda x: x[1], reverse=True))[:5]
            friends = list(sorted(
                volunteer.friend.items(), key=lambda x: x[1], reverse=True))[:5]
        ws['A4'] = 'Most visited elderly'
        ws['A5'] = 'Closest friends'
        for i in range(len(visited)):
            _ = ws.cell(row=4, column=i+2, value=str(visited[i]))
        for i in range(len(friends)):
            _ = ws.cell(row=5, column=i+2, value=str(friends[i]))
        ws.column_dimensions['A'].width = 17
        wb.save(filename)

    def output_attendances(self, filename):
        wb = Workbook()
        ws = wb.active
        ws.delete_cols(1, 56)
        ws['A1'] = 'Name'
        ws['B1'] = 'Total'
        ws.column_dimensions['A'].width = 24
        dates = []
        for v in self.all_volunteers:
            for d in v.attendance:
                if d not in dates:
                    dates.append(d)
        for i in range(len(dates)):
            _ = ws.cell(
                row=1, column=i+3, value=dates[i].date().strftime('%d/%m/%y'))
            for j in range(len(self.all_volunteers)):
                _ = ws.cell(row=j+2, column=1,
                            value=self.all_volunteers[j].get_name())
                _ = ws.cell(row=j+2, column=2,
                            value=len(self.all_volunteers[j].attendance))
                if dates[i] in self.all_volunteers[j].attendance:
                    _ = ws.cell(row=j+2, column=i+3, value=1)
                elif dates[i] in self.all_volunteers[j].pulled_out:
                    _ = ws.cell(row=j+2, column=i+3, value=0)
        wb.save(filename)

    def update_availability(self, filename):
        data = read.read_list(filename)
        if len(list(data.values())[-1][0]) != 1:
            raise ValueError('Invalid availability file')
        for name, d in data.items():
            for i in d[0]:
                self.all_volunteers[self.all_volunteers.index(
                    name)].availability.clear()
                if i == 'None of the above':
                    continue
                elif type(i) == str:
                    dates = i.split(sep=', ')
                    for date in dates:
                        date = datetime.strptime(
                            date, '%d %b').replace(year=datetime.now().year)
                        self.all_volunteers[self.all_volunteers.index(
                            name)].add_avail(date)
                else:
                    self.all_volunteers[self.all_volunteers.index(
                        name)].add_avail(i)

    def update_from_session(self, session):
        session.pulled_out = list(
            map(lambda x: self.get_person(x), session.pulled_out))
        for v in session.available_volunteers:
            self.all_volunteers[self.all_volunteers.index(
                v.get_name())].availability.remove(session.date)
            if v.get_name() not in session.pulled_out:
                self.all_volunteers[self.all_volunteers.index(
                    v.get_name())].attended(session.date)
            else:
                self.all_volunteers[self.all_volunteers.index(
                    v.get_name())].pull_out(session.date)
        for i in range(len(session.elderly_grouping)):
            elderly = session.elderly_grouping[i]
            volunteers = list(map(lambda x: self.get_person(
                x), session.volunteer_grouping[i]))
            for v in volunteers:
                for i in volunteers:
                    if i.get_name() != v.get_name():
                        self.all_volunteers[self.all_volunteers.index(v.get_name())].add_friend(
                            i.get_name())
                for e in elderly:
                    self.all_volunteers[self.all_volunteers.index(v.get_name())].visits(
                        e.get_name())

    def update_from_elderlylist(self, filename):
        """completely reassigns self.all_elderly (since we do not keep impt data in Elder objects) and updates volunteers' data for those removed elderly. """
        result = []
        data = read.read_list(filename)
        if len(list(data.values())[-1][0]) != 8:
            raise ValueError('Invalid elderly namelist')
        removed = list(filter(lambda x: x.get_name()
                              not in data.keys(), self.all_elderly))
        for i in removed:
            self.remove_elderly(i)
        for name, d in data.items():
            for blk, floor, unit, hosp, dialect, nickname, note, suspended in d:
                curr = people.Elder(name, blk, floor, unit, hosp)
                if dialect:
                    curr.set_dialect(dialect)
                if nickname:
                    curr.set_nickname(nickname)
                if note:
                    curr.add_comment(note)
                if suspended:
                    curr.suspended = True
            result.append(curr)
        self.all_elderly = result

    def update_from_volnlist(self, filename):
        """ Add/init/remove new volunteers from volnlist"""
        data = read.read_list(filename)
        if len(list(data.values())[-1][0]) != 1:
            raise ValueError('Invalid volunteer list')
        for name, d in data.items():
            if name not in self.all_volunteers:
                curr = people.Volunteer(name)
                for dialect in d[0]:
                    if dialect:
                        curr.set_dialect(dialect)
                self.all_volunteers.append(curr)
        removed = list(filter(lambda x: x.get_name()
                              not in data.keys(), self.all_volunteers))
        for i in removed:
            self.remove_volunteer(i.get_name())

    def output_availability_summary(self, filename):
        dates = []
        for v in self.all_volunteers:
            for date in v.availability:
                if date not in dates:
                    dates.append(date)
        dates.sort()
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Name'
        ws.column_dimensions['A'].width = 24
        count = 0
        for i in range(len(dates)):
            _ = ws.cell(row=1, column=i + 2,
                        value=dates[i].strftime('%d/%m/%y'))
            for j in range(len(self.all_volunteers)+1):
                if j == len(self.all_volunteers):
                    _ = ws.cell(row=j+2, column=1, value='Total')
                    _ = ws.cell(row=j+2, column=i+2, value=count)
                    count = 0
                    continue
                _ = ws.cell(row=j+2, column=1,
                            value=self.all_volunteers[j].get_name())
                if dates[i] in self.all_volunteers[j].availability:
                    _ = ws.cell(row=j+2, column=i+2, value=1)
                    count += 1
        wb.save(filename)

    def end_year(self):
        newpath = os.path.join('Archive', str(
            datetime.today().year) + ' Year end output')
        if not os.path.exists(newpath):
            os.makedirs(newpath)
            os.makedirs(os.path.join(newpath, 'All profiles'))
        self.output_attendances(os.path.join(newpath, str(
            datetime.today().year) + ' attendance.xlsx'))
        for v in self.all_volunteers:
            self.output_volunteer_detail(v.get_name(),
                                         os.path.join(newpath, 'All profiles', v.get_name() + ' profile.xlsx'))
            v.attendance.clear()
            v.availability.clear()
            v.pulled_out.clear()
            for i in v.visited:
                v.visited[i] = math.ceil(v.visited[i]/2)
            for j in v.friend:
                v.friend[j] = math.ceil(v.friend[j]/2)

    def cancel_session(self, date):
        for v in self.all_volunteers:
            if date in v.availability:
                v.availability.remove(date)

    # Helper methods
    def create_elderly_grouping(self, n):
        """ Create n groups of elderly with similar total hospitability score, output in a list of lists containing ELders. """
        result = []
        d = {}
        for i in self.all_elderly:
            if not i.suspended:
                d[i.get_name()] = i.get_hosp()
        groups = binpacking.to_constant_bin_number(d, n)
        for j in groups:
            curr = []
            for elder in j.keys():
                curr.append(self.get_person(elder))
            result.append(curr)
        result = list(map(lambda x: self.sort_elderly(x), result))
        return result

    def get_person(self, query):
        """ Returns the Elder/Volunteer object given str query """
        if self.all_volunteers.count(query) > 1 or self.all_elderly.count(query) > 1:
            raise KeyError(
                f'There are more than one {query}s, be more specific.')
        elif query in self.all_elderly:
            return self.all_elderly[self.all_elderly.index(query)]
        elif query in self.all_volunteers:
            return self.all_volunteers[self.all_volunteers.index(query)]
        else:
            raise KeyError(f'{query} is not recognised.')

    def fam_score(self, volunteer, group):
        """
        Volunteer -> volunteer obj, group -> list of elderly
        Returns int score of how often the volunteer has visited all elderly in the group
        """
        score = 0
        for i in group:
            score += volunteer.times_visited(i.get_name())
            if i.dialect and volunteer.dialect and i.dialect in volunteer.dialect:
                score += 5
        return score

    def sort_elderly(self, group):
        """ Sort elderly group by floor then by blk in ascending order """
        return sorted(sorted(group, key=lambda x: x.floor), key=lambda x: x.blk)

    def volunteer_pref(self, volunteer, elderly_grouping):
        return list(map(lambda x: self.fam_score(volunteer, x), elderly_grouping))

    def get_available(self, date):
        result = list(
            filter(lambda x: x.is_available(date), self.all_volunteers))
        if result:
            return result
        else:
            date = date.strftime('%d/%m/%Y')
            raise KeyError(f'Invalid date; Nobody is available on {date}')

    def load_elderly_grouping(self, filename):
        result = []
        ws = load_workbook(filename=filename).active
        if 'group' not in ws['B1'].value.lower():
            raise KeyError('Invalid elderly grouping file.')
        for i in ws['B']:
            if i.value == None:
                curr_group = self.sort_elderly(curr_group)
                result.append(curr_group)
                continue
            elif 'group' in i.value.lower():
                curr_group = []
                continue
            if '(' in i.value:
                i = i.value[:i.value.index('(')]
                i = i.strip()
                elderly = self.get_person(i)
            else:
                elderly = self.get_person(i.value.strip())
            if elderly.suspended:
                raise KeyError(f'{elderly} is suspended.')
            curr_group.append(elderly)
        curr_group = self.sort_elderly(curr_group)
        result.append(curr_group)
        return result

    def friend_list(self, volunteer, volunteer_data):
        result = list(volunteer.friend.items())
        for v in volunteer_data:
            if v.get_name() not in volunteer.friend and v.get_name() != volunteer.get_name():
                result.append((v.get_name(), 0))
        result = sorted(result, key=lambda x: x[0])
        result = sorted(result, key=lambda x: x[1], reverse=True)
        return result


class Session(Planner):
    """ Contains the detailing for one specific sunday and able to write to excel. """

    def __init__(self, date, elderly_grouping, available_volunteers):
        self.date = date
        self.elderly_grouping = elderly_grouping
        self.available_volunteers = available_volunteers
        self.pulled_out = []
        self.volunteer_grouping = []

    # Main methods
    def write_detailing(self, filename):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Detailing'
        wb.create_sheet('Friend Data')
        ws.delete_cols(2, 11)
        bold = Font(bold=True)
        ws['B2'] = 'Name'
        ws['C2'] = 'Blk'
        ws['D2'] = 'Floor'
        ws['E2'] = 'Unit'
        ws['F2'] = 'Notes'
        ws['H2'] = self.date.date().strftime('%d/%m/%Y')
        ws['H2'].font = bold
        for col in ws.iter_cols(min_col=2, max_col=6, min_row=2, max_row=2):
            for cell in col:
                cell.font = bold
        ws.column_dimensions['F'].width = 33
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['H'].width = 11
        ws.column_dimensions['I'].width = 17
        ws.column_dimensions['K'].width = 24
        wb.save(filename)
        self.write_elderly(filename)
        self.write_volunteer(filename)
        self.write_friend_data(filename)

    def complete(self):
        self.volunteer_grouping = self.load_volunteer_grouping(
            'Sunday detailing.xlsx')
        self.archive()

    def pull_out(self, volunteer):
        """ Remove 1 volunteer at a time """
        self.pulled_out.append(
            self.available_volunteers[self.available_volunteers.index(volunteer)])
        self.available_volunteers.pop(
            self.available_volunteers.index(volunteer))

    def un_pull_out(self, volunteer):
        """ Undo remove volunteer 1 at a time """
        self.available_volunteers.append(
            self.pulled_out[self.pulled_out.index(volunteer)])
        self.pulled_out.pop(self.pulled_out.index(volunteer))

    def archive(self):
        wb = load_workbook('Sunday detailing.xlsx')
        del wb['Friend Data']
        ws = wb['Detailing']
        ws.delete_cols(11, 2)
        target = './Archive/Previous Detailing/' + \
            self.date.date().strftime('%d-%m-%Y') + '.xlsx'
        wb.save(target)

    # Helper methods
    def load_volunteer_grouping(self, filename):
        wb = load_workbook(filename=filename)
        ws = wb['Detailing']
        result = []
        group_rows = list(map(lambda x: x.row, filter(
            lambda x: x.value != None and 'Group' in x.value, ws['H'])))
        for row in range(3, len(ws['I'])+1):
            if row in group_rows[1:]:
                result.append(curr)
            if row in group_rows:
                curr = []
            if ws.cell(row=row, column=9).value in self.pulled_out:
                raise KeyError(
                    f'{ws.cell(row=row, column=9).value} has been pulled out')
            elif ws.cell(row=row, column=9).value:
                curr.append(ws.cell(row=row, column=9).value.strip())
        result.append(curr)
        return result

    def write_elderly(self, filename):
        row = 3
        wb = load_workbook(filename=filename)
        ws = wb['Detailing']
        grp = 0
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
        for group in self.elderly_grouping:
            grp += 1
            _ = ws.cell(row=row, column=8, value='Group ' + str(grp) + ':')
            _.border = border
            for e in group:
                for col in range(2, 7):
                    if col == 2:
                        name = e.get_name()
                        if e.nickname:
                            name = name + ' (' + e.nickname + ')'
                        _ = ws.cell(row=row, column=col, value=name)
                        _.alignment = Alignment(wrap_text=True)
                    elif col == 3:
                        _ = ws.cell(row=row, column=col, value=e.blk)
                    elif col == 4:
                        _ = ws.cell(row=row, column=col, value=e.floor)
                    elif col == 5:
                        _ = ws.cell(row=row, column=col, value=e.unit)
                    elif col == 6:
                        _ = ws.cell(row=row, column=col)
                        _.alignment = Alignment(wrap_text=True)
                        if e.note:
                            _ = ws.cell(row=row, column=col, value=e.note)
                        row += 1
                    _.border = border
            row += 1

        wb.save(filename)

    def write_volunteer(self, filename):

        def is_prime(n):
            for i in range(2, int(n**0.5)+1):
                if n % i == 0:
                    return False
            return True

        wb = load_workbook(filename=filename)
        ws = wb['Detailing']
        row = 3
        for v in self.available_volunteers:
            for col in range(11, 13):
                if col == 11:
                    _ = ws.cell(row=row, column=col, value=v.get_name())
                    if is_prime(len(v.attendance)+1) and (1+len(v.attendance)) > 3:
                        _.font = Font(underline='single')
                    if (len(v.pulled_out)+len(v.attendance)) > 3:
                        if v.pull_out_rate() >= 0.45:
                            _.fill = PatternFill('solid', fgColor='FF0000')
                        elif v.pull_out_rate() > 0.3:
                            _.fill = PatternFill('solid', fgColor='FC6969')
                        elif v.pull_out_rate() > 0.2:
                            _.fill = PatternFill('solid', fgColor='FFA7A7')
                elif col == 12:
                    _ = ws.cell(row=row, column=col, value=str(
                        self.volunteer_pref(v, self.elderly_grouping)))
                    row += 1
        wb.save(filename)

    def write_friend_data(self, filename):
        """ Top and last 25% friends (max 5) amongst available_volunteers """
        k = min(len(self.available_volunteers)//4, 5)
        d = {}
        for v in self.available_volunteers:
            d[v.get_name()] = self.friend_list(v, self.available_volunteers)
        wb = load_workbook(filename=filename)
        ws = wb['Friend Data']
        ws.delete_cols(1, 12)
        ws['A1'] = 'Name'
        ws['B1'] = f'Top {k} friends'
        _ = ws.cell(row=1, column=(2+k), value=f'Last {k} friends')
        row = 2
        for name, lst in d.items():
            topk = lst[:k]
            lastk = lst[-k:]
            _ = ws.cell(row=row, column=1, value=name)
            col = 2
            for i in range(k):
                x = ws.cell(row=row, column=col, value=str(topk[0]))
                y = ws.cell(row=row, column=col+k, value=str(lastk[0]))
                topk.pop(0)
                lastk.pop(0)
                col += 1
            row += 1
        ws.auto_filter.ref = ws.dimensions
        ws.column_dimensions['A'].width = 24
        for i in range(2, 2*k+2):
            ws.column_dimensions[get_column_letter(i)].width = 27
        wb.save(filename)
