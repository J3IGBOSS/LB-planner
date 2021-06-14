import read
import store
import people
from planner import *
from openpyxl import load_workbook
from datetime import datetime
from tkinter import *
from tkinter import filedialog

# This is the script I used to manually enter the data before to start the program.


def init_voln(filename):
    result = []
    data = read.read_list(filename)
    for name, d in data.items():
        if name:
            name = name.strip()
            curr = people.Volunteer(name)
            for dialect in d[0]:
                if dialect:
                    curr.set_dialect(dialect)
            result.append(curr)
    return result


def init_elderly(filename):
    result = []
    data = read.read_list(filename)
    for name, d in data.items():
        for blk, floor, unit, hosp, dialect, nickname, note in d:
            curr = people.Elder(name, blk, floor, unit, hosp)
            if dialect:
                curr.set_dialect(dialect)
            if nickname:
                curr.set_nickname(nickname)
            if note:
                curr.add_comment(note)
        result.append(curr)
    return result


def map_compile(filename, elderly_data, volunteer_data):
    ws = load_workbook(filename).active
    for group in ws.iter_rows(values_only=True):
        elderly = list(group[:10])
        volunteers = list(group[10:])
        while None in elderly:
            elderly.remove(None)
        while None in volunteers:
            volunteers.remove(None)
        for v in volunteers:
            if v:
                v = v.strip()
                for i in volunteers:
                    if i != v and i:
                        i = i.strip()
                        volunteer_data[volunteer_data.index(v)].add_friend(
                            volunteer_data[volunteer_data.index(i)].get_name())
                for e in elderly:
                    if e:
                        if '(' in e:
                            e = e[:e.index('(')]
                        e = e.strip()
                        volunteer_data[volunteer_data.index(v)].visits(
                            elderly_data[elderly_data.index(e)].get_name())


def map_avail(filename, elderly_data, volunteer_data):
    data = read.read_list(filename)
    for name, d in data.items():
        for i in d[0]:
            if i == 'None of the above':
                continue
            elif type(i) == str:
                dates = i.split(sep=', ')
                for date in dates:
                    date = datetime.strptime(
                        date, '%d %b').replace(year=datetime.now().year)
                    volunteer_data[volunteer_data.index(name)].add_avail(date)
            else:
                volunteer_data[volunteer_data.index(name)].add_avail(i)


def map_attendance(filename, elderly_data, volunteer_data):
    data = read.read_list(filename)
    dates = load_workbook(filename).active[1][1:]
    for name, d in data.items():
        for i in range(len(dates)):
            if d[0][i] == 0:
                volunteer_data[volunteer_data.index(
                    name)].pull_out(dates[i].value)
            elif d[0][i] == 1:
                volunteer_data[volunteer_data.index(
                    name)].attended(dates[i].value)


all_elderly = init_elderly('./init/Elderly namelist.xlsx')
all_volunteers = init_voln('./init/Volunteer namelist.xlsx')
map_compile('./init/Compile.xlsx', all_elderly, all_volunteers)
# map_avail('test/avail_test.xlsx', all_elderly, all_volunteers)
map_attendance('./init/attendance.xlsx', all_elderly, all_volunteers)


LB = Planner(all_elderly, all_volunteers)
LB.remove_volunteer('Leon')
x = ['low kam lan', 'yeo shaw buan', 'siew yong hoe']
for i in x:
    LB.suspend(i)
SUNDAY = None

store.save_pkl(LB, 'planner_data.pkl')
# store.save_pkl(SUNDAY, 'session_data.pkl')
