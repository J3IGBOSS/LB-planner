from openpyxl import load_workbook


def read_list(file):
    """ Takes .xlsx file with colA=name and row 1 detailing the name of data. Returns {name:[data1,data2...], ...}  """
    wb = load_workbook(filename=file)
    result = {}
    for data in wb.active.iter_rows(min_row=2, values_only=True):
        result[data[0]] = [data[1:]]
    while None in result:
        del result[None]
    return result
